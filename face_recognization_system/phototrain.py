import cv2
import openpyxl
import os
import numpy as np
from openpyxl import Workbook
from PIL import Image
import matplotlib.pyplot as plt

class FaceRecognitionTrainer:
    def __init__(self, model_path, excel_path, photo_folder):
        self.clf = cv2.face.LBPHFaceRecognizer_create()
        self.model_path = model_path
        self.excel_file = excel_path
        self.photo_folder = photo_folder
        self.init_excel_file()
        self.student_data = self.load_student_data(self.excel_file)
        self.face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
        self.fig, self.ax = plt.subplots()  # Initialize the plot once

    def init_excel_file(self):
        if not os.path.exists(self.excel_file):
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Student Data"
            columns = ["ID", "Name", "Division", "Gender", "DOB", "Email", "Phone No", "Address", "Teacher", "PhotoSample"]
            for col_num, column_title in enumerate(columns, 1):
                sheet.cell(row=1, column=col_num, value=column_title)
            workbook.save(self.excel_file)
            print(f"Excel file created at {self.excel_file}")

    def load_student_data(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active
        student_data = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            id = int(row[0])  # Ensure ID is read as integer
            name = row[1]
            photo_sample = row[-1]  # Assuming last column is PhotoSample
            student_data[id] = (name, photo_sample)
        # Logging to verify data loading
        print("Loaded student data:")
        for id, data in student_data.items():
            print(f"ID: {id}, Name: {data[0]}, Photo: {data[1]}")
        return student_data

    def show_image(self, img_path):
        img = cv2.imread(img_path, cv2.IMREAD_GRAYSCALE)
        self.ax.clear()  # Clear the previous image
        self.ax.imshow(img, cmap='gray')
        self.ax.set_title(f"Training Image: {img_path}")
        plt.pause(1)  # Pause to allow the image to be displayed

    def train_model(self):
        face_samples = []
        ids = []

        for id, data in self.student_data.items():
            name, photo_sample = data
            # Generate path template and load all images for each student
            photo_dir = os.path.join(self.photo_folder, str(id))
            for img_num in range(100):
                img_path = os.path.join(photo_dir, f"{id}_{name}_{img_num}.jpg")
                if os.path.exists(img_path):
                    self.show_image(img_path)  # Display the training image
                    pil_image = Image.open(img_path).convert('L')  # Convert to grayscale
                    img_numpy = np.array(pil_image, 'uint8')
                    faces = self.face_cascade.detectMultiScale(img_numpy)
                    if len(faces) == 0:
                        print(f"No face detected in {img_path}")
                    for (x, y, w, h) in faces:
                        face_samples.append(img_numpy[y:y+h, x:x+w])
                        ids.append(id)
                        print(f"Added face of ID {id} from {img_path}")
                else:
                    print(f"Image path {img_path} does not exist.")

        if face_samples and ids:
            self.clf.train(face_samples, np.array(ids))
            self.clf.save(self.model_path)
            print("Model trained and saved as", self.model_path)
        else:
            print("No valid image data found for training.")

if __name__ == "__main__":
    base_dir = os.path.dirname(os.path.abspath(__file__))
    model_path = os.path.join(base_dir, "trainer.yml")
    excel_path = os.path.join(base_dir, "student_data.xlsx")
    photo_folder = os.path.join(base_dir, "photos")
    
    try:
        frt = FaceRecognitionTrainer(model_path, excel_path, photo_folder)
        plt.ion()  # Turn on interactive mode for matplotlib
        frt.train_model()
        plt.ioff()  # Turn off interactive mode
        plt.show()  # Ensure the last image is displayed
    except FileNotFoundError as e:
        print(e)
    except cv2.error as e:
        print(f"OpenCV error: {e}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
