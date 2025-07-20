import cv2
import openpyxl
import os
import numpy as np
from openpyxl import Workbook
from PIL import Image
import matplotlib.pyplot as plt

class FaceRecognitionTrainer:
    def __init__(self, model_path, excel_path, photo_folder, test_split=0.2):
        self.clf = cv2.face.LBPHFaceRecognizer_create()
        self.model_path = model_path
        self.excel_file = excel_path
        self.photo_folder = photo_folder
        self.test_split = test_split  # Percentage of data to use for testing
        self.init_excel_file()
        self.student_data = self.load_student_data(self.excel_file)
        self.face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
        self.fig, self.ax = plt.subplots()  # Initialize the plot once

    def init_excel_file(self):
        if not os.path.exists(self.excel_file):
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Student Data"
            columns = ["ID", "Name", "Division", "Gender", "DOB", "Email", "Phone No", "Address", "Teacher", "PhotoSample"]
            for col_num, column_title in enumerate(columns, 1):
                sheet.cell(row=1, column=col_num, value=column_title)
            workbook.save(self.excel_file)
            print(f"Excel file created at {self.excel_file}")

    def load_student_data(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active
        student_data = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            id = int(row[0])  # Ensure ID is read as integer
            name = row[1]
            photo_sample = row[-1]  # Assuming last column is PhotoSample
            student_data[id] = (name, photo_sample)
        # Logging to verify data loading
        print("Loaded student data:")
        for id, data in student_data.items():
            print(f"ID: {id}, Name: {data[0]}, Photo: {data[1]}")
        return student_data

    def show_image(self, img_path):
        img = cv2.imread(img_path, cv2.IMREAD_GRAYSCALE)
        self.ax.clear()  # Clear the previous image
        self.ax.imshow(img, cmap='gray')
        self.ax.set_title(f"Training Image: {img_path}")
        plt.pause(1)  # Pause to allow the image to be displayed

    def split_data(self, ids, photo_dir):
        data = [(id, os.path.join(photo_dir, f"{id}_{name}_{img_num}.jpg")) 
                for id, name in ids.items() for img_num in range(100) 
                if os.path.exists(os.path.join(photo_dir, f"{id}_{name}_{img_num}.jpg"))]
        np.random.shuffle(data)
        split_idx = int(len(data) * (1 - self.test_split))
        return data[:split_idx], data[split_idx:]

    def train_model(self):
        face_samples = []
        ids = []
        total_images = 0
        used_images = 0
        test_images = 0

        for id, data in self.student_data.items():
            name, photo_sample = data
            # Generate path template and load all images for each student
            photo_dir = os.path.join(self.photo_folder, str(id))
            if not os.path.isdir(photo_dir):
                print(f"Directory {photo_dir} does not exist.")
                continue

            train_data, test_data = self.split_data({id: name}, photo_dir)
            total_images += len(train_data) + len(test_data)
            test_images += len(test_data)

            for _, img_path in train_data:
                print(f"Processing {img_path}...")
                self.show_image(img_path)  # Display the training image
                pil_image = Image.open(img_path).convert('L')  # Convert to grayscale
                img_numpy = np.array(pil_image, 'uint8')
                faces = self.face_cascade.detectMultiScale(img_numpy)
                if len(faces) == 0:
                    print(f"No face detected in {img_path}")
                for (x, y, w, h) in faces:
                    face_samples.append(img_numpy[y:y+h, x:x+w])
                    ids.append(id)
                    used_images += 1
                    print(f"Added face of ID {id} from {img_path}")

        if total_images > 0:
            train_data_percent = (used_images / total_images) * 100
            test_data_percent = (test_images / total_images) * 100
            print(f"Training data usage: {used_images}/{total_images} images ({train_data_percent:.2f}%)")
            print(f"Testing data usage: {test_images}/{total_images} images ({test_data_percent:.2f}%)")

        if face_samples and ids:
            self.clf.train(face_samples, np.array(ids))
            self.clf.save(self.model_path)
            print("Model trained and saved as", self.model_path)
        else:
            print("No valid image data found for training.")

    def predict(self, img_path):
        pil_image = Image.open(img_path).convert('L')  # Convert to grayscale
        img_numpy = np.array(pil_image, 'uint8')
        faces = self.face_cascade.detectMultiScale(img_numpy)
        if len(faces) == 0:
            print(f"No face detected in {img_path}")
            return []

        predictions = []
        for (x, y, w, h) in faces:
            id, confidence = self.clf.predict(img_numpy[y:y+h, x:x+w])
            predictions.append((id, confidence))
            print(f"Predicted ID {id} with confidence {confidence} for {img_path}")
        return predictions

    def evaluate_model(self):
        total_predictions = 0
        correct_predictions = 0

        for id, data in self.student_data.items():
            name, photo_sample = data
            photo_dir = os.path.join(self.photo_folder, str(id))
            if not os.path.isdir(photo_dir):
                print(f"Directory {photo_dir} does not exist.")
                continue

            _, test_data = self.split_data({id: name}, photo_dir)
            for _, img_path in test_data:
                predictions = self.predict(img_path)
                for pred_id, confidence in predictions:
                    total_predictions += 1
                    if pred_id == id:
                        correct_predictions += 1

        accuracy = (correct_predictions / total_predictions) * 100 if total_predictions > 0 else 0
        print(f"Model evaluation completed. Accuracy: {accuracy:.2f}%")

if __name__ == "__main__":
    model_path = "trainerdata.yml"  # Ensure this path is correct
    excel_path = r"E:\face_recognization_system\student_data.xlsx"
    photo_folder = r"E:\face_recognization_system\photos"
    
    try:
        frt = FaceRecognitionTrainer(model_path, excel_path, photo_folder, test_split=0.2)
        plt.ion()  # Turn on interactive mode for matplotlib
        frt.train_model()
        frt.evaluate_model()  # Evaluate the model after training
        plt.ioff()  # Turn off interactive mode
        plt.show()  # Ensure the last image is displayed
    except FileNotFoundError as e:
        print(e)
    except cv2.error as e:
        print(f"OpenCV error: {e}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
