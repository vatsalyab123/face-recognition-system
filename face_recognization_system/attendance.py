import cv2
import openpyxl
import numpy as np
import os
from datetime import datetime, timedelta
from tkinter import Tk, messagebox, Label, Button

class FaceRecognitionSystem:
    def __init__(self, model_path, excel_path):
        self.clf = cv2.face.LBPHFaceRecognizer_create()
        self.read_model(model_path)
        self.excel_file = excel_path
        self.init_excel_file()
        self.student_data = self.load_student_data(self.excel_file)
        self.face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
        self.attendance_logged = False
        self.recognized_face_id = None
        self.recognized_face_name = None
        self.cancelled_by_user = False

    def read_model(self, model_path):
        if not os.path.exists(model_path):
            raise FileNotFoundError(f"The model file {model_path} does not exist.")
        self.clf.read(model_path)

    def init_excel_file(self):
        if not os.path.exists(self.excel_file):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Student Data"
            columns = ["ID", "Name", "Division", "Gender", "DOB", "Email", "Phone No", "Address", "Teacher", "PhotoSample"]
            for col_num, column_title in enumerate(columns, 1):
                sheet.cell(row=1, column=col_num, value=column_title)
            workbook.save(self.excel_file)

    def load_student_data(self, excel_path):
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active
        student_data = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            id = int(row[0])  # Ensure ID is read as integer
            name = row[1]
            student_data[id] = name
        # Logging to verify data loading
        print("Loaded student data:")
        for id, name in student_data.items():
            print(f"ID: {id}, Name: {name}")
        return student_data

    def create_attendance_sheet(self):
        attendance_file = "attendance.xlsx"
        if not os.path.exists(attendance_file):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Attendance"
            columns = ["ID", "Name", "Date", "Period", "Time"]
            for col_num, column_title in enumerate(columns, 1):
                sheet.cell(row=1, column=col_num, value=column_title)
            workbook.save(attendance_file)
        return attendance_file

    def get_current_period(self, time):
        periods = {
            "1st Period": ("00:00", "00:50"),
            "2nd Period": ("01:00", "01:50"),
            "3rd Period": ("02:00", "02:50"),
            "4th Period": ("03:00", "03:50"),
            "5th Period": ("04:00", "04:50"),
            "6th Period": ("05:00", "05:50"),
            "7th Period": ("06:00", "06:50"),
            "8th Period": ("07:00", "07:50"),
            "9th Period": ("08:00", "08:50"),
            "10th Period": ("09:00", "09:50"),
            "11th Period": ("10:00", "10:50"),
            "12th Period": ("11:00", "11:50"),
            "13th Period": ("12:00", "12:50"),
            "14th Period": ("13:00", "13:50"),
            "15th Period": ("14:00", "14:50"),
            "16th Period": ("15:00", "15:50"),
            "17th Period": ("16:00", "16:50"),
            "18th Period": ("17:00", "17:50"),
            "19th Period": ("18:00", "18:50"),
            "20th Period": ("19:00", "19:50"),
            "21st Period": ("20:00", "20:50"),
            "22nd Period": ("21:00", "21:50"),
            "23rd Period": ("22:00", "22:50"),
            "24th Period": ("23:00", "23:50"),
        }

        for period, (start, end) in periods.items():
            if start <= time <= end:
                return period
        return "Unknown"

    
    def recognize_faces(self):
        self.attendance_logged = False  # Reset attendance flag
        self.cancelled_by_user = False  # Reset cancelled flag
        while not self.attendance_logged and not self.cancelled_by_user:
            self.recognize_faces_once()

    def recognize_faces_once(self):
        cap = cv2.VideoCapture(0)
        start_time = datetime.now()
        while (datetime.now() - start_time) < timedelta(seconds=20):
            if self.cancelled_by_user:
                break  # Exit the loop if user cancels recognition
            ret, frame = cap.read()
            if not ret:
                print("Failed to grab frame")
                break

            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = self.face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(30, 30))

            for (x, y, w, h) in faces:
                face = gray[y:y+h, x:x+w]
                id, confidence = self.clf.predict(face)
                
                # Logging recognized ID and confidence
                print(f"Recognized ID: {id}, Confidence: {confidence}")

                name = self.student_data.get(id, "Unknown")
                confidence_text = f"{round(100 - confidence)}%"
                text = f"ID: {id}, Name: {name}, Confidence: {confidence_text}"

                if confidence < 50:  # High confidence
                    cv2.putText(frame, text, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)
                    cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 255, 0), 2)
                elif 50 <= confidence < 80:  # Moderate confidence
                    cv2.putText(frame, text, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 255), 2)
                    cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 255, 255), 2)

                if not self.attendance_logged and not self.cancelled_by_user:
                    self.recognized_face_id = id
                    self.recognized_face_name = name

            cv2.imshow("Face Recognition", frame)
            if cv2.waitKey(1) == 13:  # Press 'Enter' to break the loop
                break

        cap.release()
        cv2.destroyAllWindows()

        # After 20 seconds, ask for confirmation if a face was recognized
        if self.recognized_face_id and not self.attendance_logged:
            if not self.show_confirmation_dialog(self.recognized_face_id, self.recognized_face_name):
                print("Recognition canceled by the user")

    def show_confirmation_dialog(self, id, name):
        result = messagebox.askyesnocancel("Confirm Attendance", f"Is this {name}?")
        if result is None:
        # Cancel pressed, return False to indicate cancellation
            self.cancelled_by_user = True
            return False
        elif result:
        # Yes pressed, log attendance
            self.log_attendance(id, name)
            return True
        else:
        # No pressed, return False to indicate reattempt
            self.recognized_face_id = None
            self.recognized_face_name = None
            return False

    def log_attendance(self, id, name):
        attendance_file = self.create_attendance_sheet()
        wb = openpyxl.load_workbook(attendance_file)
        sheet = wb.active
        current_time = datetime.now()
        date = current_time.strftime("%Y-%m-%d")
        time = current_time.strftime("%H:%M:%S")
        period = self.get_current_period(current_time.strftime("%H:%M"))

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == id and row[2] == date and row[3] == period:
                print(f"Attendance for ID {id} has already been logged for {period} on {date}.")
                return

        new_row = [id, name, date, period, time]
        sheet.append(new_row)
        wb.save(attendance_file)
        self.attendance_logged = True
        print(f"Logged attendance for ID {id}, Name: {name}, Date: {date}, Period: {period}, Time: {time}")
        print("Attendance successfully taken")
        self.quit_gui()  # Quit the GUI after logging attendance

    def run_gui(self):
        self.root = Tk()
        self.root.title("Face Recognition Attendance System")

        def on_take_attendance():
            self.recognize_faces()

        label = Label(self.root, text="Face Recognition Attendance System", font=("Arial", 16))
        label.pack(pady=20)

        take_attendance_button = Button(self.root, text="Take Attendance", command=on_take_attendance, font=("Arial", 14))
        take_attendance_button.pack(pady=20)

        self.root.mainloop()

    def quit_gui(self):
        self.root.quit()
        self.root.destroy()

if __name__ == "__main__":
    model_path = "trainer.yml"  # Ensure this path is correct
    excel_path = r"E:\face_recognization_system\student_data.xlsx"

    try:
        frs = FaceRecognitionSystem(model_path, excel_path)
        frs.run_gui()
    except FileNotFoundError as e:
        print(e)
    except cv2.error as e:
        print(f"OpenCV error: {e}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
