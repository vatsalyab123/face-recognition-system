import tkinter as tk
from tkinter import ttk, Label, LabelFrame, Frame, Button, StringVar, messagebox
from PIL import Image, ImageTk
from tkinter.constants import RIDGE, W
from openpyxl import Workbook, load_workbook
import os
import cv2
import re
import time

class FaceRecognitionSystem:
    def __init__(self, root):
        self.root = root
        self.root.title("Face Recognition System")
        self.root.geometry('1261x843+0+0')

        self.course = StringVar()
        self.year = StringVar()
        self.semester = StringVar()
        self.name = StringVar()
        self.division = StringVar()
        self.id = StringVar()
        self.gender = StringVar()
        self.dob = StringVar()
        self.email = StringVar()
        self.phone_no = StringVar()
        self.address = StringVar()
        self.teacher = StringVar()

        # Initialize Excel file
        self.excel_file = 'student_data.xlsx'
        self.init_excel_file()

        # Load and display images
        base_dir = os.path.dirname(os.path.abspath(__file__))
        image_path = os.path.join(base_dir, 'college_images', 'stanford.jpeg')
        self.photoImg1 = self.load_image(image_path, (400, 104))
        self.photoImg2 = self.load_image(image_path, (400, 104))
        self.photoImg3 = self.load_image(image_path, (461, 104))

        # Image labels
        f1_lbl = Label(self.root, image=self.photoImg1)
        f1_lbl.place(x=0, y=0)
        f2_lbl = Label(self.root, image=self.photoImg2)
        f2_lbl.place(x=400, y=0)
        f3_lbl = Label(self.root, image=self.photoImg3)
        f3_lbl.place(x=800, y=0)

        # Title label
        title_lb = Label(self.root, text="STUDENT ATTENDANCE MANAGEMENT SYSTEM SOFTWARE",
                         font=("times new roman", 16, "bold"), bg="white", fg="black")
        title_lb.place(x=0, y=104, width=1261, height=30)

        # Left Frame for student details
        left_frame = LabelFrame(self.root, bd=2, bg='white', relief=RIDGE, text="Student Details",
                                font=("times new roman", 12, "bold"))
        left_frame.place(x=5, y=140, width=620, height=490)

        # Current Course Frame
        current_course_frame = LabelFrame(left_frame, bd=2, bg='white', relief=RIDGE, text="Current Course",
                                          font=("times new roman", 12, "bold"))
        current_course_frame.place(x=5, y=0, width=610, height=125)
        self.create_course_comboboxes(current_course_frame)

        # Class Student Frame
        class_student_frame = LabelFrame(left_frame, bd=2, bg='white', relief=RIDGE, text="Class Student Information",
                                         font=("times new roman", 12, "bold"))
        class_student_frame.place(x=5, y=125, width=610, height=335)
        self.create_student_info_entries(class_student_frame)

        # Right Frame for additional student details
        right_label_frame = LabelFrame(self.root, bd=2, bg='white', relief=RIDGE, text="Student Details",
                                       font=("times new roman", 12, "bold"))
        right_label_frame.place(x=640, y=140, width=620, height=490)

        # Search Frame
        search_frame = LabelFrame(right_label_frame, bd=2, bg="white", relief=RIDGE, text="Search System",
                                  font=("times new roman", 15, "bold"))
        search_frame.place(x=5, y=5, width=610, height=70)

        search_label = Label(search_frame, text="Search By:", font=("times new roman", 8, "bold"), bg="red")
        search_label.grid(row=0, column=0, padx=10, pady=5, sticky=W)

        self.search_combo = ttk.Combobox(search_frame, font=("times new roman", 8, "bold"), state="readonly", width=20)
        self.search_combo["values"] = ("Select", "ID", "Phone_No")
        self.search_combo.current(0)
        self.search_combo.grid(row=0, column=1, padx=2, pady=10, sticky=W)

        self.search_entry = ttk.Entry(search_frame, width=20, font=("times new roman", 8, "bold"))
        self.search_entry.grid(row=0, column=2, padx=10, pady=5, sticky=W)

        search_btn = Button(search_frame, text="Search", width=14, font=("times new roman", 7, "bold"), bg="blue", fg="white", command=self.search_data)
        search_btn.grid(row=0, column=3, padx=10)

        showAll_btn = Button(search_frame, text="Show All", width=14, font=("times new roman", 7, "bold"), bg="blue", fg="white", command=self.load_data)
        showAll_btn.grid(row=0, column=4, padx=10)

        # Table Frame
        table_frame = Frame(right_label_frame, bd=2, bg="white", relief=RIDGE)
        table_frame.place(x=5, y=80, width=610, height=390)
        self.create_table(table_frame)

    def load_image(self, path, size):
        img = Image.open(path)
        img = img.resize(size, Image.LANCZOS)
        return ImageTk.PhotoImage(img)

    def create_course_comboboxes(self, frame):
        labels_texts = [("Department", ["Select Department", "AI&DS", "Computer", "IT", "Civil", "Mechanical"]),
                        ("Course", ["Select Course", "DS", "ML", "AI", "SOFTWARE", "RESOURCE"]),
                        ("Year", ["Select Year", "2020-21", "2021-22", "2022-23", "2023-24", "2024-25"]),
                        ("Semester", ["Select Semester", "1-SEM", "2-SEM", "3-SEM", "4-SEM", "5-SEM", "6-SEM", "7-SEM", "8-SEM"])]
        for i, (label_text, values) in enumerate(labels_texts):
            row, col = divmod(i, 2)
            Label(frame, text=label_text, font=("times new roman", 12, "bold")).grid(row=row, column=col * 2, padx=10, pady=5, sticky=W)
            combo = ttk.Combobox(frame, font=("times new roman", 12, "bold"), state="readonly", values=values)
            combo.current(0)
            combo.grid(row=row, column=col * 2 + 1, padx=10, pady=5, sticky=W)
            if label_text == "Department":
                self.department_combo = combo
            elif label_text == "Course":
                self.course_combo = combo
            elif label_text == "Year":
                self.year_combo = combo
            elif label_text == "Semester":
                self.semester_combo = combo

    def create_student_info_entries(self, frame):
        labels_texts = [
            ("Student ID", self.id), ("Student Name", self.name), ("Class Division", self.division),
            ("Gender", self.gender), ("DOB", self.dob), ("Email", self.email),
            ("Phone No", self.phone_no), ("Address", self.address), ("Teacher Name", self.teacher)
        ]

        for i, (label_text, variable) in enumerate(labels_texts):
            row, col = divmod(i, 2)
            Label(frame, text=label_text, font=("Times New Roman", 12, "bold")).grid(row=row, column=col * 2, padx=10, pady=5, sticky=W)

            if label_text == "Gender":
                gender_combo = ttk.Combobox(frame, textvariable=variable, font=("Times New Roman", 12, "bold"), state="readonly", values=["Male", "Female", "Other"])
                gender_combo.current(0)
                gender_combo.grid(row=row, column=col * 2 + 1, padx=10, pady=5, sticky=W)
                self.gender_combo = gender_combo
            else:
                entry = ttk.Entry(frame, textvariable=variable, width=20, font=("Times New Roman", 12, "bold"))
                entry.grid(row=row, column=col * 2 + 1, padx=10, pady=5, sticky=W)

        btn_frame = Frame(frame, bd=2, relief=RIDGE, bg="white")
        btn_frame.place(x=0, y=200, width=610, height=35)

        Button(btn_frame, text="Save", command=self.add_data, width=16, font=("Times New Roman", 12, "bold"), bg="blue", fg="white").grid(row=0, column=0)
        Button(btn_frame, text="Update", command=self.update_data, width=16, font=("Times New Roman", 12, "bold"), bg="blue", fg="white").grid(row=0, column=1)
        Button(btn_frame, text="Delete", command=self.delete_data, width=16, font=("Times New Roman", 12, "bold"), bg="blue", fg="white").grid(row=0, column=2)
        Button(btn_frame, text="Reset", command=self.reset_data, width=16, font=("Times New Roman", 12, "bold"), bg="blue", fg="white").grid(row=0, column=3)

        btn_frame1 = Frame(frame, bd=2, relief=RIDGE, bg="white")
        btn_frame1.place(x=0, y=235, width=610, height=35)

        Button(btn_frame1, text="Take Photo Sample", command=self.take_photo_sample, width=32, font=("Times New Roman", 12, "bold"), bg="blue", fg="white").grid(row=0, column=0)
        
    def create_table(self, frame):
        scroll_x = ttk.Scrollbar(frame, orient="horizontal")
        scroll_y = ttk.Scrollbar(frame, orient="vertical")

        self.student_table = ttk.Treeview(frame, column=("ID", "Name", "Division", "Gender", "DOB", "Email", "Phone No", "Address", "Teacher", "PhotoSample"), xscrollcommand=scroll_x.set, yscrollcommand=scroll_y.set)

        scroll_x.pack(side="bottom", fill="x")
        scroll_y.pack(side="right", fill="y")

        scroll_x.config(command=self.student_table.xview)
        scroll_y.config(command=self.student_table.yview)

        self.student_table.heading("ID", text="Student ID")
        self.student_table.heading("Name", text="Student Name")
        self.student_table.heading("Division", text="Division")
        self.student_table.heading("Gender", text="Gender")
        self.student_table.heading("DOB", text="DOB")
        self.student_table.heading("Email", text="Email")
        self.student_table.heading("Phone No", text="Phone No")
        self.student_table.heading("Address", text="Address")
        self.student_table.heading("Teacher", text="Teacher")
        self.student_table.heading("PhotoSample", text="PhotoSample")

        self.student_table["show"] = "headings"

        for col in self.student_table["columns"]:
            self.student_table.column(col, width=100)

        self.student_table.pack(fill="both", expand=1)
        self.student_table.bind("<ButtonRelease>", self.get_cursor)
        self.load_data()

    def init_excel_file(self):
        if not os.path.exists(self.excel_file):
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Student Data"
            columns = ["ID", "Name", "Division", "Gender", "DOB", "Email", "Phone No", "Address", "Teacher", "PhotoSample"]
            for col_num, column_title in enumerate(columns, 1):
                sheet.cell(row=1, column=col_num, value=column_title)
            workbook.save(self.excel_file)


    def add_data(self):
        if self.id.get() == "" or self.name.get() == "" or self.course_combo.get() == "Select Course" or self.department_combo.get() == "Select Department":
            messagebox.showerror("Error", "All fields are required!")
            return

        try:
            workbook = load_workbook(self.excel_file)
            sheet = workbook.active
            student_id = self.id.get()
            existing_ids = [sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row + 1)]
            if student_id in existing_ids:
                messagebox.showerror("Error", "Student ID already exists!")
                return

        # Placeholder for photo_path
            photo_path = ""

        # Create a placeholder row to append to the sheet
            row_to_append = [
                self.id.get(), self.name.get(), self.division.get(), self.gender.get(),
                self.dob.get(), self.email.get(), self.phone_no.get(),
                self.address.get(), self.teacher.get(), photo_path
            ]

        # Append the row to the sheet
            sheet.append(row_to_append)

        # Save the workbook
            workbook.save(self.excel_file)

        # Call method to update photo_path in the excel file
            self.update_excel_photo_path(student_id, photo_path)

            messagebox.showinfo("Success", "Student added successfully!")
            self.load_data()

        except Exception as e:
            messagebox.showerror("Error", f"Error due to: {str(e)}")

            

    def update_data(self):
        if self.id.get() == "":
            messagebox.showerror("Error", "Student ID must be provided for update!")
            return

        try:
            workbook = load_workbook(self.excel_file)
            sheet = workbook.active
            student_id = self.id.get()
            for i in range(2, sheet.max_row + 1):
                if sheet.cell(row=i, column=1).value == student_id:
                    sheet.cell(row=i, column=2, value=self.name.get())
                    sheet.cell(row=i, column=3, value=self.division.get())
                    sheet.cell(row=i, column=4, value=self.gender.get())
                    sheet.cell(row=i, column=5, value=self.dob.get())
                    sheet.cell(row=i, column=6, value=self.email.get())
                    sheet.cell(row=i, column=7, value=self.phone_no.get())
                    sheet.cell(row=i, column=8, value=self.address.get())
                    sheet.cell(row=i, column=9, value=self.teacher.get())
                    workbook.save(self.excel_file)
                    messagebox.showinfo("Success", "Student data updated successfully!")
                    self.load_data()
                    return
            messagebox.showerror("Error", "Student ID not found!")
        except Exception as e:
            messagebox.showerror("Error", f"Error due to: {str(e)}")

    def delete_data(self):
        if self.id.get() == "":
            messagebox.showerror("Error", "Student ID must be provided for deletion!")
            return

        try:
            workbook = load_workbook(self.excel_file)
            sheet = workbook.active
            student_id = self.id.get()
            for i in range(2, sheet.max_row + 1):
                if sheet.cell(row=i, column=1).value == student_id:
                    sheet.delete_rows(i)
                    workbook.save(self.excel_file)
                    messagebox.showinfo("Success", "Student data deleted successfully!")
                    self.load_data()
                    return
            messagebox.showerror("Error", "Student ID not found!")
        except Exception as e:
            messagebox.showerror("Error", f"Error due to: {str(e)}")

    def reset_data(self):
        self.id.set("")
        self.name.set("")
        self.division.set("")
        self.gender.set("")
        self.dob.set("")
        self.email.set("")
        self.phone_no.set("")
        self.address.set("")
        self.teacher.set("")
        self.department_combo.current(0)
        self.course_combo.current(0)
        self.year_combo.current(0)
        self.semester_combo.current(0)

    def load_data(self):
        try:
            workbook = load_workbook(self.excel_file)
            sheet = workbook.active
            self.student_table.delete(*self.student_table.get_children())
            for i in range(2, sheet.max_row + 1):
                self.student_table.insert("", "end", values=(
                    sheet.cell(row=i, column=1).value, sheet.cell(row=i, column=2).value,
                    sheet.cell(row=i, column=3).value, sheet.cell(row=i, column=4).value,
                    sheet.cell(row=i, column=5).value, sheet.cell(row=i, column=6).value,
                    sheet.cell(row=i, column=7).value, sheet.cell(row=i, column=8).value,
                    sheet.cell(row=i, column=9).value, sheet.cell(row=i, column=10).value  # Display photo path
                ))
        except Exception as e:
            messagebox.showerror("Error", f"Error due to: {str(e)}")


    def get_cursor(self, event):
        cursor_row = self.student_table.focus()
        content = self.student_table.item(cursor_row)
        data = content['values']
        self.id.set(data[0])
        self.name.set(data[1])
        self.division.set(data[2])
        self.gender.set(data[3])
        self.dob.set(data[4])
        self.email.set(data[5])
        self.phone_no.set(data[6])
        self.address.set(data[7])
        self.teacher.set(data[8])

    def search_data(self):
        search_by = self.search_combo.get()
        search_term = self.search_entry.get()

        if search_by == "Select":
            messagebox.showerror("Error", "Please select a search criteria!")
            return
        if search_term == "":
            messagebox.showerror("Error", "Please enter a search term!")
            return

        try:
            workbook = load_workbook(self.excel_file)
            sheet = workbook.active
            self.student_table.delete(*self.student_table.get_children())
            for i in range(2, sheet.max_row + 1):
                if (search_by == "ID" and sheet.cell(row=i, column=1).value == search_term) or \
                   (search_by == "Phone_No" and sheet.cell(row=i, column=7).value == search_term):
                    self.student_table.insert("", "end", values=(
                        sheet.cell(row=i, column=1).value, sheet.cell(row=i, column=2).value,
                        sheet.cell(row=i, column=3).value, sheet.cell(row=i, column=4).value,
                        sheet.cell(row=i, column=5).value, sheet.cell(row=i, column=6).value,
                        sheet.cell(row=i, column=7).value, sheet.cell(row=i, column=8).value, 
                        sheet.cell(row=i, column=9).value,
                        sheet.cell(row=i, column=10).value
                    ))
        except Exception as e:
            messagebox.showerror("Error", f"Error due to: {str(e)}")





    def take_photo_sample(self):
        student_id = self.id.get()
        student_name = self.name.get()
        if not student_id or not student_name:
            messagebox.showerror("Error", "Please enter the Student ID and Name.")
            return

        cap = cv2.VideoCapture(0)
        if not cap.isOpened():
            messagebox.showerror("Error", "Unable to open the camera.")
            return

        sample_count = 0
        start_time = time.time()
        photo_dir = os.path.join("photos", student_id)
        if not os.path.exists(photo_dir):
            os.makedirs(photo_dir)

        while sample_count < 100 and time.time() - start_time < 30:  # Capture 100 samples within 30 seconds
            ret, frame = cap.read()
            if not ret:
                messagebox.showerror("Error", "Failed to capture image.")
                break

            photo_path = os.path.join(photo_dir, f"{student_id}_{student_name}_{sample_count}.jpg")
            cv2.imwrite(photo_path, frame)
            sample_count += 1

        # Show the frame in a window
            cv2.imshow("Capturing Photos", frame)

        # Handle window events
            if cv2.waitKey(1) & 0xFF == ord('q'):
                break

        cap.release()
        cv2.destroyAllWindows()

        self.update_excel_photo_path(student_id, photo_path)
        messagebox.showinfo("Info", f"Captured {sample_count} photos for Student ID: {student_id}.")

    

    def update_excel_photo_path(self, student_id, photo_path, update=False):
    # Load the workbook and select the active worksheet
        wb = load_workbook(self.excel_file)
        ws = wb.active

    # Column index for the photo path (10th column)
        photo_column_index = 10  # 10th column index is 10 (1-based index)

    # Iterate through the rows to find the student ID
        for row in ws.iter_rows(min_row=2, max_col=10):
            if row[0].value == student_id:  # Assuming student ID is in the 1st column (index 0)
                current_photo_paths = row[photo_column_index - 1].value  # Adjusted for 0-based index

            # Update or replace the photo path
                if current_photo_paths:
                    if update:
                        row[photo_column_index - 1].value = f"{current_photo_paths};{photo_path}"
                    else:
                        row[photo_column_index - 1].value = photo_path
                else:
                    row[photo_column_index - 1].value = photo_path
                break
        else:
            messagebox.showerror("Error", f"Student ID {student_id} not found in the Excel sheet.")

    # Save and close the workbook
        wb.save(self.excel_file)
        wb.close()


if __name__ == "__main__":
    root = tk.Tk()
    app = FaceRecognitionSystem(root)
    root.mainloop()
